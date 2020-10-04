from openpyxl import load_workbook


#class Excel_utilities():

    # @staticmethod
def getvalue(file, sheet_name, search_string, column="A"):
    wb = load_workbook(file)
    ws = wb[sheet_name]
    for row in range(1, ws.max_row + 1):
        coordinate = "{}{}".format(column, row)
        if ws[coordinate].value == search_string:
            required_row = row
            return ws.cell(row=required_row, column=2).value